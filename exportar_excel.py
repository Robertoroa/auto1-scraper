"""
Exporta los resultados del scraper a un Excel con formato visual.
"""
import json
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

RESULTS_DIR = Path(__file__).parent / "resultados"
RESULTS_DIR.mkdir(exist_ok=True)


def exportar_excel(resultados: list) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Oportunidades Auto1"

    # ── Colores ──────────────────────────────────────────────
    COLOR_HEADER   = "1A1A2E"   # azul oscuro
    COLOR_VERDE    = "C6EFCE"   # fondo fila buena (margen > 20%)
    COLOR_AMARILLO = "FFEB9C"   # margen medio (10-20%)
    COLOR_ROJO     = "FFC7CE"   # margen bajo (<10%)
    COLOR_GRIS     = "F2F2F2"   # filas alternas

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Cabecera ─────────────────────────────────────────────
    columnas = [
        ("Marca",         15),
        ("Modelo",        18),
        ("Año",            7),
        ("KM",            10),
        ("Precio Auto1",  14),
        ("Precio Mercado",15),
        ("Margen €",      12),
        ("Margen %",      11),
        ("País",           7),
        ("Canal",         10),
        ("Ref. Auto1",    14),
        ("Daños motor",   14),
        ("Revisión",      11),
        ("Link Auto1",    40),
    ]

    for col, (titulo, ancho) in enumerate(columnas, 1):
        cell = ws.cell(row=1, column=col, value=titulo)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor=COLOR_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = ancho

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    # ── Filas de datos ────────────────────────────────────────
    for i, c in enumerate(resultados, 2):
        precio_auto1  = c.get("precio_auto1", 0) or 0
        precio_mercado = c.get("precio_mercado", 0) or 0
        margen_eur = round(precio_mercado - precio_auto1) if precio_mercado else 0
        margen_pct = round((margen_eur / precio_auto1) * 100, 1) if precio_auto1 else 0

        ref = c.get("referencia", "")
        link = f"https://www.auto1.com/es/app/merchant/car/{ref}" if ref else ""

        fila = [
            c.get("marca", ""),
            c.get("modelo", ""),
            c.get("año", ""),
            c.get("km", 0),
            precio_auto1,
            precio_mercado if precio_mercado else "—",
            margen_eur if precio_mercado else "—",
            f"{margen_pct}%" if precio_mercado else "—",
            c.get("pais", ""),
            c.get("canal", ""),
            ref,
            "Sí" if c.get("danos_motor") else ("No" if c.get("danos_motor") is False else "N/D"),
            "⚠️ Sí" if c.get("requiere_revision") else "OK",
            link,
        ]

        # Color de fondo según margen
        if precio_mercado and precio_auto1:
            if margen_pct >= 20:
                bg = COLOR_VERDE
            elif margen_pct >= 10:
                bg = COLOR_AMARILLO
            else:
                bg = COLOR_ROJO
        else:
            bg = COLOR_GRIS if i % 2 == 0 else "FFFFFF"

        fill = PatternFill("solid", fgColor=bg)

        for col, valor in enumerate(fila, 1):
            cell = ws.cell(row=i, column=col, value=valor)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=False)

            # Formato número para precio y km
            if col in (4,):  # km
                cell.number_format = "#,##0"
            if col in (5, 6, 7):  # precios y margen €
                cell.number_format = "#,##0 €"

        ws.row_dimensions[i].height = 20

    # ── Totales ───────────────────────────────────────────────
    fila_total = len(resultados) + 2
    ws.cell(row=fila_total, column=1, value=f"Total: {len(resultados)} coches").font = Font(bold=True)

    # ── Metadatos ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Info")
    ws2["A1"] = "Generado"
    ws2["B1"] = datetime.now().strftime("%d/%m/%Y %H:%M")
    ws2["A2"] = "Coches"
    ws2["B2"] = len(resultados)
    ws2["A3"] = "Fuente"
    ws2["B3"] = "Auto1 B2B Dealer Portal"

    # ── Guardar ───────────────────────────────────────────────
    nombre = f"auto1_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    ruta = RESULTS_DIR / nombre
    wb.save(ruta)
    return ruta


if __name__ == "__main__":
    # Test con datos de ejemplo
    from scraper import ejecutar_scraping
    resultados = ejecutar_scraping()
    if resultados:
        ruta = exportar_excel(resultados)
        print(f"\n✅ Excel guardado: {ruta}")
        import subprocess
        subprocess.Popen(["open", str(ruta)])
    else:
        print("No hay resultados para exportar.")
